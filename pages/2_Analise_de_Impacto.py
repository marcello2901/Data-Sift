# -*- coding: utf-8 -*-
"""
Análise de Impacto dos Resultados — DataSift
============================================

Avalia o impacto de repetir conjuntos de amostras (no mínimo 3 por teste). Cada
teste é um bloco com o seu Erro Total Máximo (ETM) e o seu Intervalo de
Referência (IR) — puxados automaticamente da base de dados pelo nome do teste —
e a sua própria tabela de amostras (código de barras, Resultado 1, Resultado 2).
É possível analisar vários testes ao mesmo tempo.

Para cada amostra decide se a diferença entre os dois resultados tem impacto
analítico (excede o ETM) e/ou clínico (muda a interpretação pelo IR).

Base de dados (colunas): Teste, Equipamento, ETM, IR.

Página do app DataSift (pasta ``pages/``). Também roda de forma independente com
``streamlit run pages/2_Analise_de_Impacto.py``.
"""

import io
import os
import re

import numpy as np
import pandas as pd
import streamlit as st

# --------------------------------------------------------------------------- #
# Identidade visual (mesma paleta do DataSift)
# --------------------------------------------------------------------------- #
COLOR_PRIMARY = "#073B4C"
COLOR_SECONDARY = "#00E5FF"
COLOR_TERTIARY = "#118AB2"
COLOR_BG = "#F8F9FA"

try:
    st.set_page_config(page_title="DataSift · Impacto", page_icon="🎯", layout="wide")
except Exception:
    pass

st.markdown(
    f"""
    <style>
        .stApp {{ background-color: {COLOR_BG} !important; }}
        h1, h2, h3, h4 {{ color: {COLOR_PRIMARY} !important; font-weight: 800 !important; }}
        p, span, div[data-testid="stMarkdownContainer"], label {{ color: #212529 !important; }}
        div[data-testid="stMetricValue"] {{ color: {COLOR_PRIMARY} !important; }}
    </style>
    """,
    unsafe_allow_html=True,
)

# --------------------------------------------------------------------------- #
# Conformidade LGPD / GDPR — mesma tela do app.py; chave de sessão compartilhada
# --------------------------------------------------------------------------- #
GDPR_TERMS = """
This tool is designed to process and filter data from spreadsheets. The files you upload may contain sensitive personal data (such as full name, date of birth, national ID numbers, health information, etc.), the processing of which is regulated by data protection laws like the General Data Protection Regulation (GDPR or LGPD).

It is your sole responsibility to ensure that all data used in this tool complies with applicable data protection regulations. We strongly recommend that you only use previously anonymized data to protect the privacy of data subjects.

The responsibility for the nature of the processed data is exclusively yours.

To proceed, you must confirm that the data to be used has been properly handled and anonymized.
"""

if "lgpd_accepted" not in st.session_state:
    st.session_state.lgpd_accepted = False

if not st.session_state.lgpd_accepted:
    st.header("Terms of Use and Data Protection Compliance")
    st.markdown(GDPR_TERMS)
    accepted = st.checkbox("By checking this box, I confirm that the data provided is anonymized.")
    if st.button("Continue", type="primary", disabled=not accepted):
        st.session_state.lgpd_accepted = True
        st.rerun()
    st.stop()

APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# --------------------------------------------------------------------------- #
# Funções auxiliares
# --------------------------------------------------------------------------- #
def normalizar_serie_numerica(serie: pd.Series) -> pd.Series:
    """Converte texto/misto em float, aceitando vírgula ou ponto como decimal."""
    def _conv(x):
        if pd.isna(x):
            return np.nan
        s = str(x).strip()
        if s == "" or s.lower() in ("nan", "none", "na", "n/a", "-", "--", "."):
            return np.nan
        s = s.replace("\xa0", "").replace(" ", "")
        s = re.sub(r"[^0-9,.\-+]", "", s)
        if s in ("", "+", "-", ".", ","):
            return np.nan
        has_dot, has_comma = "." in s, "," in s
        if has_dot and has_comma:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif has_comma:
            s = s.replace(",", ".")
        if s.count(".") > 1:
            partes = s.split(".")
            s = "".join(partes[:-1]) + "." + partes[-1]
        try:
            return float(s)
        except ValueError:
            return np.nan
    return serie.apply(_conv)


def parse_ref_range(txt):
    """Extrai (limite_inferior, limite_superior) de um IR ('136 - 145', '< 190', '> 40')."""
    if pd.isna(txt):
        return (None, None)
    s = str(txt).strip()
    if s == "":
        return (None, None)
    low = s.lower()
    nums = re.findall(r"\d+(?:[.,]\d+)?", s)
    vals = [float(n.replace(",", ".")) for n in nums]
    if not vals:
        return (None, None)
    if ("<" in s or "≤" in s or "menor" in low or "até" in low or "up to" in low):
        return (None, vals[-1])
    if (">" in s or "≥" in s or "maior" in low or "acima" in low):
        return (vals[0], None)
    if len(vals) >= 2:
        return (min(vals[0], vals[1]), max(vals[0], vals[1]))
    return (None, None)


def classificar_ref(valor, limite_inf, limite_sup):
    """Classifica em Baixo / Normal / Alto; '—' se não houver intervalo."""
    if pd.isna(valor):
        return "—"
    inf_ok = limite_inf is not None and pd.notna(limite_inf)
    sup_ok = limite_sup is not None and pd.notna(limite_sup)
    if not inf_ok and not sup_ok:
        return "—"
    if inf_ok and valor < limite_inf:
        return "Baixo"
    if sup_ok and valor > limite_sup:
        return "Alto"
    return "Normal"


_COLS_ESPERADAS = ("Teste", "Equipamento", "ETM", "IR")


def _ler_tabela(conteudo: bytes, nome: str) -> pd.DataFrame:
    """
    Lê CSV ou Excel a partir dos bytes. Para CSV, testa combinações de
    separador/decimal/encoding e escolhe a primeira que traz as colunas
    esperadas — evita 'mojibake' (UTF-8 lido como latin-1) e separador errado.
    """
    nome = (nome or "").lower()
    if nome.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(conteudo), engine="openpyxl")
    tentativas = [
        dict(sep=";", decimal=",", encoding="utf-8-sig"),
        dict(sep=";", decimal=",", encoding="latin-1"),
        dict(sep=",", decimal=".", encoding="utf-8-sig"),
        dict(sep=",", decimal=".", encoding="latin-1"),
    ]
    ultimo = None
    for t in tentativas:
        try:
            df = pd.read_csv(io.BytesIO(conteudo), engine="python", **t)
        except Exception:
            continue
        cols = [str(c).strip() for c in df.columns]
        df.columns = cols
        if all(e in cols for e in _COLS_ESPERADAS):
            return df
        ultimo = df
    return ultimo


@st.cache_data(show_spinner="Lendo base de dados...")
def carregar_base(conteudo: bytes | None, nome: str | None) -> pd.DataFrame | None:
    """
    Carrega a base de testes. Prioridade: arquivo enviado pelo usuário >
    Base de Dados.csv na raiz do projeto. Devolve None se nada existir.
    """
    df = None
    if conteudo is not None:
        df = _ler_tabela(conteudo, nome)
    else:
        caminho = os.path.join(APP_DIR, "Base de Dados.csv")
        if os.path.exists(caminho):
            with open(caminho, "rb") as fh:
                df = _ler_tabela(fh.read(), "Base de Dados.csv")
    if df is None:
        return None
    df.columns = [str(c).strip() for c in df.columns]
    return df


def to_excel(df: pd.DataFrame, cols_2dec=None) -> bytes:
    """Exporta .xlsx centralizado, com autofit e 2 casas nas colunas indicadas."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    cols_2dec = set(cols_2dec or [])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Impacto")
        ws = writer.sheets["Impacto"]
        centro = Alignment(horizontal="center", vertical="center")
        n = len(df)
        for i, col in enumerate(df.columns, start=1):
            letra = get_column_letter(i)
            largura = max([len(str(col))] + [len(str(v)) for v in df[col].values])
            ws.column_dimensions[letra].width = min(largura + 2, 60)
            for r in range(1, n + 2):
                ws.cell(row=r, column=i).alignment = centro
            if col in cols_2dec:
                for r in range(2, n + 2):
                    ws.cell(row=r, column=i).number_format = "0.00"
    return output.getvalue()


def analisar_bloco(entrada: pd.DataFrame, teste, etm, ir_txt, lo, hi):
    """Valida e avalia as amostras de um teste. Devolve (df_resultado, n_validas)."""
    am = entrada.copy()
    am["Código de barras"] = am["Código de barras"].astype(str).str.strip()
    am["R1"] = normalizar_serie_numerica(am["Resultado 1"])
    am["R2"] = normalizar_serie_numerica(am["Resultado 2"])
    val = am[(am["Código de barras"] != "") & (am["Código de barras"].str.lower() != "nan")
             & am["R1"].notna() & am["R2"].notna() & (am["R1"] != 0)].reset_index(drop=True)
    if len(val) < 3:
        return None, len(val)

    val["Erro total %"] = np.abs((val["R1"] / val["R2"]) - 1) * 100
    val["Excede ETM"] = val["Erro total %"].abs() > etm if etm is not None else False
    val["Interpretação R1"] = val["R1"].apply(lambda v: classificar_ref(v, lo, hi))
    val["Interpretação R2"] = val["R2"].apply(lambda v: classificar_ref(v, lo, hi))
    val["Mudou interpretação"] = ((val["Interpretação R1"] != val["Interpretação R2"])
                                  & (val["Interpretação R1"] != "—") & (val["Interpretação R2"] != "—"))
    val["Impacto"] = np.where(val["Excede ETM"] | val["Mudou interpretação"], "Com impacto", "Sem impacto")
    val.insert(0, "Teste", teste)
    val.insert(1, "ETM (%)", etm)
    val.insert(2, "IR", ir_txt)
    return val, len(val)


def gerar_pdf(detalhe: pd.DataFrame, equipamento: str, resumo: pd.DataFrame, analise: str) -> bytes:
    """
    Gera um relatório PDF (A4 paisagem) com o 'Detalhe por amostra', o 'Resumo por
    teste' e a 'Análise crítica' do analista. Usa matplotlib (já disponível).
    """
    import textwrap
    from datetime import datetime
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    def _fmt(df):
        d = df.copy()
        if "Erro total %" in d.columns:
            d["Erro total %"] = pd.to_numeric(d["Erro total %"], errors="coerce").map(
                lambda v: "" if pd.isna(v) else f"{v:.2f}")
        for c in ("Excede ETM", "Mudou interpretação"):
            if c in d.columns:
                d[c] = d[c].map(lambda v: "Sim" if bool(v) else "Não")
        return d.astype(str)

    _rename_pdf = {"Código de barras": "Cód. barras",
                   "Interpretação R1": "Interp. R1", "Interpretação R2": "Interp. R2",
                   "Mudou interpretação": "Mudou interp."}

    def _tabela_paginas(pdf, df, titulo, subtitulo=None, por_pag=22):
        df = _fmt(df).rename(columns=_rename_pdf)
        n_pag = max(1, (len(df) + por_pag - 1) // por_pag)
        for p in range(n_pag):
            fig = plt.figure(figsize=(11.69, 8.27))          # A4 paisagem
            ax = fig.add_subplot(111); ax.axis("off")
            fig.suptitle(titulo, fontsize=15, fontweight="bold", color="#073B4C", x=0.03, ha="left")
            if subtitulo and p == 0:
                fig.text(0.03, 0.93, subtitulo, fontsize=9, color="#333333")
            chunk = df.iloc[p * por_pag:(p + 1) * por_pag]
            h = min(0.88, (len(chunk) + 1) * 0.05)           # tabela ancorada no topo
            t = ax.table(cellText=chunk.values, colLabels=list(df.columns),
                         cellLoc="center", bbox=[0.0, 0.90 - h, 1.0, h])
            t.auto_set_font_size(False); t.set_fontsize(7.5)
            for (r, c), cell in t.get_celld().items():
                cell.set_edgecolor("#CCCCCC")
                if r == 0:
                    cell.set_facecolor("#073B4C")
                    cell.set_text_props(color="white", fontweight="bold")
            pdf.savefig(fig); plt.close(fig)

    buf = io.BytesIO()
    cabecalho = f"Equipamento: {equipamento}    |    Gerado em {datetime.now():%d/%m/%Y %H:%M}"
    with PdfPages(buf) as pdf:
        _tabela_paginas(pdf, detalhe, "Análise de Impacto — Detalhe por amostra", cabecalho)
        if resumo is not None and len(resumo):
            _tabela_paginas(pdf, resumo, "Resumo por teste")
        # Página(s) da análise crítica
        texto = (analise or "").strip() or "(não preenchida)"
        linhas = []
        for par in texto.split("\n"):
            linhas.extend(textwrap.wrap(par, width=110) or [""])
        por_pag_txt = 42
        n_pag = max(1, (len(linhas) + por_pag_txt - 1) // por_pag_txt)
        for p in range(n_pag):
            fig = plt.figure(figsize=(11.69, 8.27))
            ax = fig.add_subplot(111); ax.axis("off")
            fig.suptitle("Análise crítica", fontsize=15, fontweight="bold",
                         color="#073B4C", x=0.03, ha="left")
            bloco = "\n".join(linhas[p * por_pag_txt:(p + 1) * por_pag_txt])
            fig.text(0.03, 0.90, bloco, va="top", ha="left", fontsize=10)
            pdf.savefig(fig); plt.close(fig)
    return buf.getvalue()


# =========================================================================== #
#                                INTERFACE
# =========================================================================== #
st.markdown("## 🎯 Análise de Impacto dos Resultados")
st.caption(
    "Avalia o impacto de repetir amostras (mínimo 3 por teste) — em um ou vários testes ao "
    "mesmo tempo. O Erro Total Máximo (ETM) e o Intervalo de Referência (IR) são puxados "
    "automaticamente da base de dados pelo nome do teste."
)

# ---- 1 · Base de dados ---------------------------------------------------- #
with st.container(border=True):
    st.markdown("### 1 · Base de dados de testes")
    st.caption("Colunas esperadas: **Teste**, **Equipamento**, **ETM**, **IR**. "
               "Se você não enviar um arquivo, uso o `Base de Dados.csv` do projeto.")
    arq_base = st.file_uploader("Base de dados (opcional se já houver no projeto)",
                                type=["csv", "xlsx", "xls"], key="base_uploader")

base = carregar_base(arq_base.getvalue() if arq_base is not None else None,
                     arq_base.name if arq_base is not None else None)

if base is None or base.empty:
    st.info("📄 Envie a base de dados (colunas Teste, Equipamento, ETM, IR) ou inclua "
            "o arquivo `Base de Dados.csv` na raiz do projeto para começar.")
    st.stop()

faltando = [c for c in ["Teste", "Equipamento", "ETM", "IR"] if c not in base.columns]
if faltando:
    st.error(f"A base de dados não tem a(s) coluna(s): {', '.join(faltando)}. "
             f"Colunas encontradas: {', '.join(map(str, base.columns))}.")
    st.stop()

base = base.copy()
base["_ETM_num"] = normalizar_serie_numerica(base["ETM"])
testes = sorted(base["Teste"].dropna().astype(str).str.strip().unique().tolist())
equipamentos = sorted(base["Equipamento"].dropna().astype(str).str.strip().unique().tolist())


def lookup_teste(teste):
    """Devolve (etm, ir_txt, lo, hi) da base para o teste (1ª linha correspondente)."""
    linha = base[base["Teste"].astype(str).str.strip() == str(teste)]
    etm = float(linha["_ETM_num"].dropna().iloc[0]) if not linha["_ETM_num"].dropna().empty else None
    ir_txt = str(linha["IR"].dropna().iloc[0]) if not linha["IR"].dropna().empty else ""
    lo, hi = parse_ref_range(ir_txt)
    return etm, ir_txt, lo, hi


# Estado dos blocos de teste (um id por bloco)
if "imp_blocos" not in st.session_state:
    st.session_state.imp_blocos = [1]
    st.session_state.imp_next = 2

# ---- 2 · Equipamento e testes --------------------------------------------- #
with st.container(border=True):
    st.markdown("### 2 · Equipamento e testes")
    equip_sel = st.selectbox("Equipamento (vale para todos os testes abaixo)",
                             equipamentos, index=0 if equipamentos else None)
    st.caption("Cada bloco abaixo é um **teste**, com o seu ETM e IR puxados da base e a sua "
               "própria tabela de amostras (mínimo 3). Use o ➕ da tabela para mais linhas e o "
               "botão **Adicionar teste** para mais testes.")

blocos_dados = []   # (teste, etm, ir_txt, lo, hi, entrada_df)
for bid in list(st.session_state.imp_blocos):
    with st.container(border=True):
        top = st.columns([5, 1])
        with top[0]:
            teste_sel = st.selectbox("Nome do teste", testes,
                                     index=0 if testes else None, key=f"teste_{bid}")
        with top[1]:
            st.markdown("<div style='height:1.75rem'></div>", unsafe_allow_html=True)
            if len(st.session_state.imp_blocos) > 1 and st.button("🗑️ Remover", key=f"del_{bid}"):
                st.session_state.imp_blocos.remove(bid)
                st.rerun()

        etm, ir_txt, lo, hi = lookup_teste(teste_sel)
        cE = st.columns(2)
        cE[0].metric("Erro Total Máximo (ETM)", f"{etm:.2f} %" if etm is not None else "—")
        cE[1].metric("Intervalo de Referência (IR)", ir_txt if ir_txt else "—")

        seed = pd.DataFrame({"Código de barras": ["", "", ""],
                             "Resultado 1": ["", "", ""],
                             "Resultado 2": ["", "", ""]})
        entrada = st.data_editor(
            seed, num_rows="dynamic", use_container_width=True, key=f"am_{bid}",
            column_config={
                "Código de barras": st.column_config.TextColumn("Código de barras"),
                "Resultado 1": st.column_config.TextColumn("Resultado 1"),
                "Resultado 2": st.column_config.TextColumn("Resultado 2"),
            },
        )
        blocos_dados.append((teste_sel, etm, ir_txt, lo, hi, entrada))

if st.button("➕ Adicionar teste"):
    st.session_state.imp_blocos.append(st.session_state.imp_next)
    st.session_state.imp_next += 1
    st.rerun()

# ---- Cálculo (todos os blocos) -------------------------------------------- #
partes, avisos = [], []
for teste_sel, etm, ir_txt, lo, hi, entrada in blocos_dados:
    res, q = analisar_bloco(entrada, teste_sel, etm, ir_txt, lo, hi)
    if res is None:
        if q > 0:   # bloco vazio não gera aviso; só o parcialmente preenchido (1-2)
            avisos.append((teste_sel, q))
    else:
        partes.append(res)

for teste_sel, q in avisos:
    st.warning(f"⚠️ **{teste_sel}**: informe no mínimo 3 amostras válidas (há {q}). "
               "Este teste não entrou na análise.")

if not partes:
    st.info("ℹ️ Informe pelo menos um teste com **3 amostras válidas** (código de barras + "
            "Resultado 1 e 2 numéricos, R1 ≠ 0) para ver a análise de impacto.")
    st.stop()

todos = pd.concat(partes, ignore_index=True)

# ---- 3 · Resultado da análise --------------------------------------------- #
st.markdown("### 3 · Resultado da análise de impacto")
st.caption(f"Equipamento avaliado: **{equip_sel}** · {todos['Teste'].nunique()} teste(s).")

n = len(todos)
n_etm = int(todos["Excede ETM"].sum())
n_interp = int(todos["Mudou interpretação"].sum())
n_impacto = int((todos["Impacto"] == "Com impacto").sum())

m1, m2, m3, m4 = st.columns(4)
m1.metric("Amostras analisadas", f"{n}")
m2.metric("Excedem o ETM", f"{n_etm}")
m3.metric("Mudam de interpretação", f"{n_interp}")
m4.metric("Com impacto (combinado)", f"{n_impacto}")

if n_impacto:
    st.error(f"⚠️ {n_impacto} de {n} amostra(s) **com impacto** no equipamento **{equip_sel}** "
             "(excedem o ETM e/ou mudam de interpretação). Avalie antes de liberar.")
else:
    st.success(f"✅ Nenhuma das {n} amostras apresentou impacto em **{equip_sel}**.")

# Resumo por teste
resumo = (todos.assign(_imp=(todos["Impacto"] == "Com impacto").astype(int),
                       _etm=todos["Excede ETM"].astype(int),
                       _int=todos["Mudou interpretação"].astype(int))
          .groupby("Teste")
          .agg(Amostras=("Impacto", "size"), Excedem_ETM=("_etm", "sum"),
               Mudam_interp=("_int", "sum"), Com_impacto=("_imp", "sum"))
          .reset_index())
st.markdown("**Resumo por teste**")
st.dataframe(resumo, use_container_width=True)

# Detalhe por amostra
tab = todos[["Teste", "Código de barras", "R1", "R2", "Erro total %", "Excede ETM",
             "Interpretação R1", "Interpretação R2", "Mudou interpretação", "Impacto"]].rename(
    columns={"R1": "Resultado 1", "R2": "Resultado 2"})


def _hl(v):
    return ("background-color:#FFE3E3; color:#9B1C1C; font-weight:700" if v == "Com impacto"
            else "background-color:#E7F6EC; color:#0F5132")


st.markdown("**Detalhe por amostra**")
st.dataframe(tab.style.format({"Erro total %": "{:.2f}", "Resultado 1": "{:.3f}",
                               "Resultado 2": "{:.3f}"}).map(_hl, subset=["Impacto"]),
             use_container_width=True)

# ---- 4 · Análise crítica -------------------------------------------------- #
st.markdown("### 4 · Análise crítica")
st.caption("Registre a sua avaliação dos dados acima (conclusão do analista). "
           "Este texto entra no relatório em PDF.")
analise_critica = st.text_area(
    "Análise crítica do analista", height=180, key="analise_critica",
    placeholder="Ex.: Foram avaliadas N amostras do teste X no equipamento Y. "
                "Observou-se ... . Conduta: ...",
    label_visibility="collapsed",
)

# ---- 5 · Exportar --------------------------------------------------------- #
st.markdown("### 5 · Exportar")
export = todos[["Teste", "ETM (%)", "IR", "Código de barras", "R1", "R2", "Erro total %",
                "Excede ETM", "Interpretação R1", "Interpretação R2",
                "Mudou interpretação", "Impacto"]].rename(
    columns={"R1": "Resultado 1", "R2": "Resultado 2"}).copy()
export.insert(0, "Equipamento", equip_sel)
export["Erro total %"] = export["Erro total %"].round(2)

d1, d2, d3 = st.columns(3)
with d1:
    st.download_button("⬇️ Baixar (Excel)",
                       data=to_excel(export, cols_2dec=["Erro total %", "Resultado 1", "Resultado 2"]),
                       file_name="analise_impacto.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with d2:
    csv_bytes = export.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button("⬇️ Baixar (CSV)", data=csv_bytes,
                       file_name="analise_impacto.csv", mime="text/csv")
with d3:
    st.download_button("⬇️ Baixar (PDF)",
                       data=gerar_pdf(tab, equip_sel, resumo, analise_critica),
                       file_name="analise_impacto.pdf", mime="application/pdf")
